from typing import List
from sqlalchemy.orm import Session
from app import models, schemas
from datetime import datetime
from app.config import settings
import openpyxl.writer
import openpyxl.writer.excel
import io
import openpyxl
from openpyxl.styles import PatternFill,Font, Alignment
from .. import constant 

def create_cuenta(db: Session, prestamo: schemas.Prestamo, cedula: str):
    
    # id_usuario = str(db.query(models.Usuario).filter(models.Usuario.cedula == str(cedula)).first().id)
    # print(f"==>> id_usuario: {id_usuario}")

    
    cuenta = schemas.CuentaCreate(
        prestamo_id=prestamo.id,
        moneda=prestamo.moneda,
        balance=prestamo.monto,
        usuario_cedula=cedula
    )
    db_cuenta = models.Cuenta(**cuenta.dict())

    db.add(db_cuenta)
    db.commit()
    db.refresh(db_cuenta)
    return db_cuenta


def get_cuentas(db: Session, cedula):
    #TODO
    permiso_usuario = 0
    
    print(f"==>> id_usuario: {cedula}")
    if cedula == "":
        if permiso_usuario == 0:
            cuentas_existentes = db.query(models.Cuenta).all()
            return cuentas_existentes   
        elif permiso_usuario > 0:
            cuentas_existentes = List[schemas.Cuenta]
            return cuentas_existentes
    else:   
        print(f"==>>Obteniendo todas las cuestas del usuario: {cedula}")
        cuentas_existentes = db.query(models.Cuenta).filter(models.Cuenta.usuario_cedula == cedula).all()
        return cuentas_existentes   
    
    return db.query(models.Cuenta).all()

def get_cuentaExcel(db: Session, usuario_cedula):
    #TODO
    permiso_usuario = 0

    print(f"==>>Obteniendo todas las cuentas del usuario: {usuario_cedula}")
    cuenta: schemas.CuentaExcel = db.query(models.Cuenta).filter(models.Cuenta.usuario_cedula == usuario_cedula).first()

    cuenta = calcularBalanceParaCadaMovimiento(cuenta)

    return cuenta   


def calcularBalanceParaCadaMovimiento(cuenta:schemas.CuentaExcel):
    balance = cuenta.prestamo.monto
    for index, mov in enumerate(cuenta.movimientos):
        if(mov.tipo==constant.MOVIMIENTOS_PAGO):
            balance = balance - mov.monto
            cuenta.movimientos[index].balance = balance
        elif(mov.tipo==constant.MOVIMIENTOS_INTERES):
            balance = balance + mov.monto
            cuenta.movimientos[index].balance = balance
    
    return cuenta



def get_Excel(db: Session, output, usuario_cedula):

    cuenta: schemas.CuentaExcel = get_cuentaExcel(db, usuario_cedula)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"

    horaActualReporte = datetime.now()

    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Estado de Cuenta - {settings.NOMBRE_EMPRESA}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Fecha de estado de cuenta: {horaActualReporte}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="left")

    # Nombre de cliente
    ws.merge_cells('A3:E3')
    ws['A3'] = f"Cliente: {cuenta.usuario.nombre}"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal="left")

    # Informacion de prestamo
    ws.merge_cells('A4:E4')
    ws['A4'] = f"Interes del prestamo: {cuenta.prestamo.interes_anual}%"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].alignment = Alignment(horizontal="left")

    

    ws.merge_cells('A5:E5')
    ws['A5'] = f"Monto del prestamo: {cuenta.prestamo.monto}"
    ws['A5'].font = Font(bold=True, size=12)
    ws['A5'].alignment = Alignment(horizontal="left")
    
    # Definir el color de la cabecera
    header_fill = PatternFill(start_color="00CCFF99", end_color="00CCFF99", fill_type="solid")  # Verde claro
    header_font = Font(bold=True, color="000000")  # Blanco para el texto

    # Añadir las cabeceras con fondo verde
    headers = ["Fecha", "Tipo","Detalle", "Monto", "Balance", "Realizado por"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 20  # Columna para Fecha
    ws.column_dimensions['B'].width = 16  # Columna para Detalle
    ws.column_dimensions['C'].width = 30  # Columna para Detalle
    ws.column_dimensions['D'].width = 16  # Columna para Detalle
    ws.column_dimensions['E'].width = 16  # Columna para Detalle
    ws.column_dimensions['F'].width = 20  # Columna para Detalle

    moneda_simbolo="₡"
    if(cuenta.prestamo.moneda==constant.MONEDA_DOLARES):
        moneda_simbolo="$"

    # Agregar movimientos a la tabla
    for mov in cuenta.movimientos:
        ws.append([
            mov.fecha,
            mov.tipo,
            mov.detalle,
            f"{moneda_simbolo}{mov.monto}",
            f"{moneda_simbolo}{mov.balance}",
            mov.usuarioRegistrante.nombre
        ])

    # Guardar el archivo en memoria
    wb.save(output)
    output.seek(0)
    return output