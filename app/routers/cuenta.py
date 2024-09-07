import openpyxl.writer
import openpyxl.writer.excel
from .. import models,schemas, auth
from app.crud import cuentaCRUD
from fastapi import FastAPI, Response, status, HTTPException, Depends, APIRouter
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..database import get_db
from typing import List, Optional
import io
import pandas
import openpyxl
from openpyxl.styles import PatternFill,Font, Alignment
import base64

router = APIRouter(
    prefix='/cuentas',
    tags=['Cuentas']
)

@router.post("/",status_code=status.HTTP_201_CREATED,response_model=schemas.Cuenta)
def create_cuenta(cuentaACrear: schemas.CuentaCreate, db: Session = Depends(get_db)):
    
    ## Permiso 0 = admin
    ## Permiso 1 = normal
    raise HTTPException (status_code=status.HTTP_306_RESERVED,
                        detail=f"Cuenta se crea con la creacion de un prestamo.")
    if usuarioExistente is not None:
        print(f"Usuario ya existe: {usuarioExistente}")

    cuentaACrear = cuentaCRUD.create_cuenta(db, usuario)
    print(f"Usuario creado: {nuevo_usuario}")

    return nuevo_usuario


@router.get("/",response_model=List[schemas.Cuenta])
def get_cuentas(db: Session = Depends(get_db),
                id_usuario: Optional[str] = ""):


    cuentas_existentes = cuentaCRUD.get_cuentas(db, id_usuario)
    print(f"==>> cuentas_existentes: {cuentas_existentes}")

    if not cuentas_existentes:
        raise HTTPException (status_code=status.HTTP_404_NOT_FOUND,
                            detail=f"No hay cuentas existentes.")
    
    return cuentas_existentes

''' 
@router.get("/generate-excel/",response_model=schemas.ExcelResponse)
def get_cuentas(db: Session = Depends(get_db),
                id_usuario: Optional[str] = ""):

    cuentas_existentes = cuentaCRUD.get_cuentas(db, id_usuario)[0]
    print(f"==>> cuentas_existentes: {cuentas_existentes}")

    if not cuentas_existentes:
        raise HTTPException (status_code=status.HTTP_404_NOT_FOUND,
                            detail=f"No hay cuentas existentes.")
    
    #######
    nombreSheet = "Estado de Cuenta - Prestamo"
    nombreArchivo = "EstadoDeCuenta.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombreSheet
    
    ws.append(["Name", "Age", "Occupation"])
    for name, age, occupation in zip(["Josh"], ["30"], ["Engineer"]):
        ws.append([name, age, occupation])
    
    # Save the workbook to an in-memory file
    output = io.BytesIO() 
    wb.save(output)
    output.seek(0)

    openpyxl.writer.excel.save_workbook(wb)
    # wb.save(nombreArchivo)
    print(f"==>> output: {output.getvalue()}")
    
    # Base64 encode the file content
    file_content = base64.b64encode(output.read()).decode('utf-8')

    headers = {
        'Content-Disposition': f'attachment; filename="{nombreArchivo}"'
    }

    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers, as_attachment=True)
    
''' 
@router.get("/generate-excel/")
def get_cuentas(db: Session = Depends(get_db),
                id_usuario: Optional[str] = ""):

    cuentas_existentes = cuentaCRUD.get_cuentas(db, id_usuario)[0]
    print(f"==>> cuentas_existentes: {cuentas_existentes}")

    if not cuentas_existentes:
        raise HTTPException (status_code=status.HTTP_404_NOT_FOUND,
                            detail=f"No hay cuentas existentes.")

    return get_Excel(db)


def get_Excel(db: Session = Depends(get_db)):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"

    # Nombre de tu empresa
    company_name = "Josh Le Presta"  # Nombre de tu empresa
    
    # Añadir título del documento (Nombre de la empresa y nombre del cliente)
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Estado de Cuenta - {company_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Cliente: Jane"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Definir el color de la cabecera
    header_fill = PatternFill(start_color="00CCFF99", end_color="00CCFF99", fill_type="solid")  # Verde claro
    header_font = Font(bold=True, color="FFFFFF")  # Blanco para el texto

    # Añadir las cabeceras con fondo verde
    headers = ["Fecha", "Tipo","Detalle", "Monto", "Balance", "Realizado por"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Establecer el ancho de las columnas
    ws.column_dimensions['A'].width = 28  # Columna para Fecha
    ws.column_dimensions['B'].width = 28  # Columna para Detalle

    # Agregar movimientos a la tabla
    movimientos = [{"date": "2024-09-06","tipo":"Crédito", "detail": "Pago préstamo", "amount": -100, "balance": 900, "usuario":"Jane Berrocal"}]
    for movement in movimientos:
        ws.append([
            movement["date"],
            movement["tipo"],
            movement["detail"],
            movement["amount"],
            movement["balance"],
            movement["usuario"]
        ])

    # Guardar el archivo en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar el archivo Excel como respuesta para descargar
    headers = {
        'Content-Disposition': 'attachment; filename="estado_de_cuenta.xlsx"'
    }
    
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)